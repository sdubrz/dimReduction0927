# 清洗数据
import numpy as np
from Main import DimReduce
from Main import Preprocess
import matplotlib.pyplot as plt
from sklearn.decomposition import PCA
import openpyxl
import csv
import os
from PIL import Image


def wine_quality_red():
    """红酒质量数据"""
    in_path = "E:\\Project\\result2019\\result0927\\datasets\\winequalityred\\"
    data0 = np.loadtxt(in_path+"winequality-red.csv", dtype=np.float, delimiter=";")

    (n, m) = data0.shape
    print((n, m))
    np.savetxt(in_path+"data.csv", data0[:, 0:m-1], fmt='%f', delimiter=',')
    np.savetxt(in_path+"label.csv", data0[:, m-1], fmt='%d', delimiter=',')

    print('数据清洗完毕')


def bostonHouse6912():
    path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\bostonHouse\\"
    data = np.loadtxt(path+"data.csv", dtype=np.float, delimiter=",")
    (n, m) = data.shape

    data2 = np.zeros((n, 3))
    data2[:, 0] = data[:, 5]
    data2[:, 1] = data[:, 8]
    data2[:, 2] = data[:, 12]

    np.savetxt(path+"data3.csv", data2, fmt="%f", delimiter=",")
    print("数据处理完毕")


def coil_20():
    path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\coil20obj\\"
    data = np.loadtxt(path+"data.csv", dtype=np.float, delimiter=",")
    label = np.loadtxt(path+"label.csv", dtype=np.int, delimiter=",")
    (n, m) = data.shape
    X = Preprocess.normalize(data, -1, 1)

    # dr_method = "PCA"
    # Y = DimReduce.dim_reduce(X[0:5*72, :], method=dr_method, method_k=15)
    # plt.scatter(Y[:, 0], Y[:, 1], marker='o', c=label[0:5*72])

    # Y = DimReduce.dim_reduce(X, method=dr_method, method_k=15)
    # plt.scatter(Y[:, 0], Y[:, 1], marker='o', c=label)
    # plt.title(dr_method)
    # ax = plt.gca()
    # ax.set_aspect(1)
    # plt.show()

    # 用PCA将数据压缩到 16 维
    pca = PCA(n_components=16)
    Y = pca.fit_transform(X)
    np.savetxt(path+"Y16.csv", Y, fmt='%f', delimiter=",")
    np.savetxt(path+"Y16_5class.csv", Y[0:5*72, :], fmt='%f', delimiter=",")
    print("数据处理完毕")


def news20_group():
    path = "E:\\Project\\DataLab\\20news-18828\\"
    data = np.loadtxt(path+"vector.csv", dtype=np.float, delimiter=",")

    (n, m) = data.shape
    print((n, m))
    X = Preprocess.normalize(data, -1, 1)
    pca = PCA(n_components=2)
    Y = pca.fit_transform(X)
    plt.scatter(Y[:, 0], Y[:, 1])
    plt.show()


def see_news20():
    path = "E:\\Project\\DataLab\\20news-18828\\"
    data = np.loadtxt(path + "vector64.csv", dtype=np.float, delimiter=",")

    (n, m) = data.shape
    print((n, m))
    X = Preprocess.normalize(data[0:1000, :], -1, 1)

    pca = PCA(n_components=2)
    Y = pca.fit_transform(X)

    plt.scatter(Y[:, 0], Y[:, 1])
    plt.show()


def eclipse(a, b, x0=0.0, y0=0.0):
    """
    计算椭圆
    :param a: 长轴
    :param b: 短轴
    :param x0: 中心点横坐标
    :param y0: 中心点纵坐标
    :return:
    """
    points = []
    for t in range(0, 360):
        x = a * np.cos(t*np.pi/180) + x0
        y = b * np.sin(t*np.pi/180) + y0
        points.append([x, y])

    return np.array(points)


def show_eclipse():
    eclipse1 = eclipse(6, 5)
    eclipse2 = eclipse(36, 25)

    plt.subplot(121)
    plt.plot(eclipse1[:, 0], eclipse1[:, 1])
    ax = plt.gca()
    ax.set_aspect(1)  #
    plt.title("6:5")

    plt.subplot(122)
    plt.plot(eclipse2[:, 0], eclipse2[:, 1])
    ax = plt.gca()
    ax.set_aspect(1)  #
    plt.title("36:25")

    plt.show()


def swissroll1800():
    path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\swissroll2000-3\\"
    # data = np.loadtxt(path, dtype=np.str, delimiter=",", encoding='utf-8')
    # X = data[:, :].astype(np.float)
    # print(data.shape)
    data = openpyxl.load_workbook(path+"data.xlsx")
    print(data)
    ws = data.active
    print(ws.cell(row=1, column=1).value)

    X = np.zeros((2000, 3))
    for i in range(0, 2000):
        for j in range(0, 3):
            X[i, j] = ws.cell(row=i+1, column=j+1).value

    np.savetxt(path+"data.csv", X, fmt='%f', delimiter=",")


def gapminder():
    """将gapminder数据中的异常数据进行处理"""
    path = "E:\\Project\\DataLab\\gapminder\\"
    f = open(path+"gapminder_unfiltered.csv")
    reader = csv.reader(f)
    data = list(reader)
    n = len(data)
    for i in range(0, n):
        temp_list = data[i]
        if len(temp_list) == 6:
            continue
        alter_list = []
        first_str = temp_list[0]
        for j in range(1, len(temp_list)-5):
            first_str = first_str + temp_list[j]
        alter_list.append(first_str)
        for j in range(len(temp_list)-5, len(temp_list)):
            alter_list.append(temp_list[j])
        data[i] = alter_list

    file2 = open(path+"gapminder.csv", 'w', newline='')
    file_writer = csv.writer(file2)
    for row in data:
        file_writer.writerow(row)
    file2.close()


def gapminder_years():
    """对gapminder 数据根据年份进行划分"""
    path = "E:\\Project\\DataLab\\gapminder\\"
    in_file = open(path+"gapminder_notitle.csv")
    reader = csv.reader(in_file)
    data = list(reader)
    dicts = {}
    n = len(data)
    for record in data:
        year = record[2]
        if year in dicts:
            temp_list = dicts[year]
            temp_list.append(record)
        else:
            temp_list = []
            temp_list.append(record)
            dicts[year] = temp_list

    for key in dicts:
        print(key)
        out_file = open(path+key+".csv", 'w', newline='')
        writer = csv.writer(out_file)
        key_values = dicts[key]
        for record in key_values:
            writer.writerow(record)
        out_file.close()


def gapminder_label():
    """对 gapminder 数据添加标签"""
    path = "E:\\Project\\DataLab\\gapminder\\"
    year = 1952
    label_dict = {}
    label_count = 0
    while year <= 2007:
        in_file = open(path+"years\\"+str(year)+".csv")
        reader = csv.reader(in_file)
        data = list(reader)
        n = len(data)
        label = np.zeros((n, 1))
        country = []
        out_data = []
        for i in range(0, n):
            record = data[i]
            country.append(record[0])
            if record[1] in label_dict:
                label[i, 0] = label_dict[record[1]]
            else:
                label_count += 1
                label_dict[record[1]] = label_count
                label[i, 0] = label_count
            temp_list = []
            for j in range(3, len(record)):
                temp_list.append(record[j])
            out_data.append(temp_list)
        in_file.close()

        out_path = path + "gapminder" + str(year) + "\\"
        if not os.path.exists(out_path):
            os.makedirs(out_path)
        np.savetxt(out_path+"label.csv", label, fmt='%d', delimiter=",")
        data_file = open(out_path+"data.csv", 'w', newline='')
        data_writer = csv.writer(data_file)
        for row in out_data:
            data_writer.writerow(row)
        data_file.close()
        country_file = open(out_path+"country.csv", 'w', newline='')
        country_writer = csv.writer(country_file)
        for c in country:
            country_writer.writerow(c)
        country_file.close()

        print(year)
        year += 5


def coli_custer_pca():
    """
    对coli的每个数据分别进行PCA降维
    :return:
    """
    path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\coil20obj_16_5class\\"
    data = np.loadtxt(path+"data.csv", dtype=np.float, delimiter=",")

    for i in range(0, 5):
        X = data[i*72:(i+1)*72, :]
        X = Preprocess.normalize(X)
        pca = PCA(n_components=2)
        Y = pca.fit_transform(X)
        np.savetxt(path+"pca"+str(i+1)+".csv", Y, fmt='%f', delimiter=",")
    print("finished")


def wine_quality():
    path = "E:\\Project\\DataLab\\wineQuality\\"
    red_data = np.loadtxt(path+"winequality-red.csv", dtype=np.float, delimiter=";")
    white_data = np.loadtxt(path+"winequality-white.csv", dtype=np.float, delimiter=";")
    data = np.zeros((1000, 11))
    label = np.ones((1000, 1))
    quality = np.zeros((1000, 1))

    for i in range(0, 500):
        data[i, :] = red_data[i*3, 0:11]
        label[i] = 1
        quality[i] = red_data[i*3, 11]

    for i in range(0, 500):
        data[i+500, :] = white_data[i*9, 0:11]
        label[500+i] = 2
        quality[i+500] = white_data[i*9, 11]

    np.savetxt(path+"data.csv", data, fmt='%f', delimiter=",")
    np.savetxt(path+"label.csv", label, fmt='%d', delimiter=",")
    np.savetxt(path+"quality.csv", quality, fmt='%d', delimiter=",")
    print("sampling finished")


def pose_test():
    path = "E:\\Project\\DataLab\\MoCap\\testframe\\poseTest.txt"
    reader_file = open(path, encoding='utf-8')
    pose = reader_file.read()
    print(pose)


def mnist():
    path = "E:\\Project\\DataLab\\MNIST\\"
    X = np.loadtxt(path+"data.csv", dtype=np.float, delimiter=",")
    pca = PCA(n_components=50)
    Y = pca.fit_transform(X)

    np.savetxt(path+"Y.csv", Y, fmt='%f', delimiter=",")


def mnist_50m():
    path = "E:\\Project\\DataLab\\MNIST\\"
    data = np.loadtxt(path+"data.csv", dtype=np.float, delimiter=",")
    (n, m) = data.shape
    X = data.tolist()
    label = np.loadtxt(path+"label.csv", dtype=np.int, delimiter=",")
    data_list = []
    for i in range(0, 10):
        data_list.append([])

    index = 0
    while index < n:
        data_list[label[index]].append(X[index])
        index += 1

    for i in range(0, 10):
        np.savetxt(path+str(i)+".csv", np.array(data_list[i]), fmt='%d', delimiter=",")

    print('success')


def mnist_50m_class(path1="", origin_path=""):
    """
    采样
    :param path1:
    :param origin_path:
    :return:
    """
    # path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\MNIST50mclass2\\"
    X_ = np.loadtxt(path1, dtype=np.float, delimiter=",")
    (n, m) = X_.shape
    X = X_.tolist()
    origin_ = np.loadtxt(origin_path, dtype=np.int, delimiter=",")
    origin = origin_.tolist()

    small_data = []  # PCA之后的数据
    small_1 = []  # 原始的数据
    for i in range(0, n):
        if i % 30 == 0:
            small_data.append(X[i])
            small_1.append(origin[i])

    return np.array(small_data), np.array(small_1)
    # np.savetxt(path+"small.csv", np.array(small_data), fmt='%f', delimiter=",")
    # np.savetxt(path+"origin.csv", np.array(small_1), fmt='%d', delimiter=",")


def mnist_pictures(path=""):
    """生成MNIST的图片"""
    # path = "E:\\Project\\DataLab\\MNIST\\"
    # path = "E:\\Project\\result2019\\result1026without_straighten\\datasets\\MNIST50mclass2_874\\"
    X = np.loadtxt(path+"origin.csv", dtype=np.int, delimiter=",")
    (n, m) = X.shape
    X = 255*np.ones(X.shape) - X
    # X = np.maximum(X, 1)
    # label = np.loadtxt(path+"label.csv", dtype=np.int, delimiter=",")
    # count = np.zeros((10, 1))
    picture_path = path + "pictures\\"
    if not os.path.exists(picture_path):
        os.makedirs(picture_path)

    for i in range(0, n):
        new_data = np.reshape(X[i, :], (28, 28))
        im = Image.fromarray(new_data.astype(np.uint8))
        # plt.imshow(new_data, cmap=plt.cm.gray, interpolation='nearest')
        # im.show()
        # im.save(path+"pictures\\"+str(label[i])+"\\"+str(int(count[label[i]]))+".png")
        im.save(picture_path+str(i)+".png")
        # count[label[i]] += 1
        if i % 1000 == 0:
            print(i)

    print("finished")


def mnist_run():
    """ 处理 MNIST数据 """
    path = "E:\\Project\\DataLab\\MNIST50m\\"
    origin_path = "E:\\Project\\DataLab\\MNIST\\"

    for digit in range(0, 10):
        print("digit = ", digit)
        X, origin_X = mnist_50m_class(path+str(digit)+".csv", origin_path+str(digit)+".csv")
        (n, m) = X.shape
        label = np.ones((n, 1)) * digit

        digit_path = path + "MNIST50mclass" + str(digit) + "_" + str(n) + "\\"
        if not os.path.exists(digit_path):
            os.makedirs(digit_path)
        np.savetxt(digit_path+"data.csv", X, fmt='%f', delimiter=",")
        np.savetxt(digit_path+"origin.csv", origin_X, fmt='%d', delimiter=",")
        np.savetxt(digit_path+"label.csv", label, fmt='%d', delimiter=",")
        mnist_pictures(digit_path)


def mocap_teen():
    """
    对人体骨骼数据删除一些关节，制作一个简化版的数据
    :return:
    """
    read_path = "E:\\Project\\DataLab\\MoCap\\cleanData\\swimmer.csv"
    write_path = "E:\\Project\\DataLab\\MoCap\\mocap_teen\\swimmer.csv"
    data = np.loadtxt(read_path, dtype=np.float, delimiter=",")
    (n, m) = data.shape  # 目前的数据中共有28个关节，是84维的数据

    left_points = [2, 3, 4, 7, 8, 9, 11, 14, 16, 17, 18, 19, 24, 25, 26]  # 要保留的关节号，是从1开始计数的
    X = np.zeros((n, len(left_points)*3))
    for i in range(0, len(left_points)):
        j = left_points[i] - 1
        X[:, 3*i] = data[:, j*3]
        X[:, 3*i+1] = data[:, j*3+1]
        X[:, 3*i+2] = data[:, j*3+2]

    np.savetxt(write_path, X, fmt='%f', delimiter=",")


def mocap_less_frame():
    """
    减少人体骨骼数据的帧数
    :return:
    """
    read_path = "E:\\Project\\DataLab\\MoCap\\mocap_teen\\swimmer.csv"
    write_path = "E:\\Project\\DataLab\\MoCap\\mocap_teen\\swimmer_small.csv"
    data = np.loadtxt(read_path, dtype=np.float, delimiter=",")
    (n, m) = data.shape
    data_list = data.tolist()
    X = []

    for i in range(0, n):
        if i % 10 == 0:
            X.append(data_list[i])

    np.savetxt(write_path, np.array(X), fmt='%f', delimiter=",")


def mnist_pca_research():
    path = "E:\\Project\\DataLab\\MNIST\\"
    data = np.loadtxt(path + "data.csv", dtype=np.float, delimiter=",")
    (n, m) = data.shape
    label = np.loadtxt(path + "label.csv", dtype=np.int, delimiter=",")

    pca = PCA(n_components=m)
    pca.fit(data)
    vectors = pca.components_  # 所有的特征向量
    values = pca.explained_variance_  # 所有的特征值
    np.savetxt(path+"eigenvectors.csv", vectors, fmt='%f', delimiter=",")
    np.savetxt(path+"eigenvalues.csv", values, fmt='%f', delimiter=",")
    print("finished")


if __name__ == '__main__':
    # wine_quality_red()
    # bostonHouse6912()
    # coil_20()
    # news20_group()
    # see_news20()
    # show_eclipse()
    # swissroll1800()
    # gapminder()
    # gapminder_label()
    # coli_custer_pca()
    # wine_quality()
    # pose_test()
    # mnist_50m_class()
    # mnist_pictures()
    # mnist_50m()
    mnist_run()
    # mocap_teen()
    # mocap_less_frame()
    # mnist_pca_research()

